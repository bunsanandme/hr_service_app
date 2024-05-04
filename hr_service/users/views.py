from django.db.models.query import QuerySet
from django.forms import FileInput, Textarea
from django.shortcuts import get_object_or_404, redirect, render
from phonenumber_field.formfields import PhoneNumberField
from django.http import HttpResponse
from django.urls import reverse, reverse_lazy
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic import CreateView, UpdateView, DeleteView
from django.contrib.auth.hashers import make_password
from django.contrib.messages.views import SuccessMessageMixin
from django.core.files.storage import FileSystemStorage


from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

from .models import User
from .forms import UserForm

from openpyxl import load_workbook
from datetime import datetime
import xlwt


@method_decorator(login_required, name='dispatch')
class UserListView(ListView):
    model = User
    context_object_name = "user"
    template_name = "user_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Список сотрудников"
        return context
    
    def get_queryset(self):
        print(self.request.GET)
        queryset = User.objects.exclude(id=self.request.user.id)
        return queryset


@method_decorator(login_required, name='dispatch')   
class UserDetailView(DetailView):
    model = User
    template_name = "user_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Профиль сотрудника"
        return context


@method_decorator(login_required, name='dispatch')
class UserCreateView(CreateView):
    model = User
    form_class = UserForm
    template_name = 'user_create.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Создание сотрудника"
        return context
    
    def form_valid(self, form):       
        instance = form.save(commit=False)
        instance.user = self.request.user
        instance.save() 

        return redirect("user-update", id=instance.pk)


@method_decorator(login_required, name='dispatch')
class UserUpdateView(SuccessMessageMixin, UpdateView):
    model = User
    form_class = UserForm
    pk_url_kwarg = 'id'
    template_name = 'user_update.html'
    success_message = "Данные сотрудника изменены!"
    
    def get_form(self, form_class=None):
        form = super(UserUpdateView, self).get_form(form_class)
        form.fields['document_scan'].widget = FileInput()
        form.fields['photo'].widget = FileInput()
        form.fields['education'].widget = Textarea()
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Редактирование"
        return context

    def form_valid(self, form):
        user = User.objects.get(id = self.kwargs["id"])
        user.document_scan = form.cleaned_data['document_scan']
        user.save()
        if self.request.POST["changed_password"]:
            pswd = self.request.POST["changed_password"]
            form.instance.password = make_password(pswd)
        super().form_valid(form)
        pk = self.kwargs["id"]
        return super(UserUpdateView, self).form_valid(form)
    
    def get_success_url(self) -> str:
        pk = self.kwargs["id"]
        return reverse("user-detail", kwargs={"pk": pk})


@method_decorator(login_required, name='dispatch')
class UserDeleteView(DeleteView):
    model = User

    def get(self, request, *args, **kwargs):
        obj = get_object_or_404(User, id=self.kwargs.get('id'))
        obj.delete()
        return redirect('user-list')


def export_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="export_{datetime.today().strftime("%Y-%m-%d")}.xlsx"'
    wb = xlwt.Workbook(encoding='cp1251')
    ws = wb.add_sheet('Users')
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Username', 'First name', 'Last name', 'Email address', "Position", "Education", "Start Job Date"]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()
    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email', "position", "education", "start_job_date")
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)
    return response

def import_excel(request):
    context_dict = {"title": "Импорт данных"}
    if request.method == 'POST':
        if request.FILES['excel_file']:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                new_user = User(last_name=row[0],first_name=row[1],username=row[2],email=row[3],education=row[4],position=row[5])
                new_user.save() 
    return render(request, "import_excel.html", context=context_dict)